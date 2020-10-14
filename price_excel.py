import pandas as pd
import glob
import openpyxl as px
import win32com.client
import time
import pythoncom
import glob
import mouse_action
import main
import os
def byd_import():
    #基本データのファイル全取得
    #files = glob.glob('\\hi-nas\事務所\商品DB\商品登録データ\本DB取り込み(情報システム作業用)\完了/**/*')
    files = glob.glob('E:\\src\\Python\\price_excel\\venv\\excel\*.xlsx')
    
    print('ファイル検索完了['+str(len(files))+'件]登録処理開始')    
    #Excelマクロの呼び出し
    pythoncom.CoInitialize()
    app = win32com.client.Dispatch("Excel.Application")
    #Excelを表示させる
    app.Visible = False
     #Excelマクロ開く
    wb_macro = app.Workbooks.Open(Filename=r"E:\\src\\Python\\price_excel\\venv\\excel\\Excelツール\\TST【CH-MK】Excelツール_02.価格アップロード_V1.0.xlsm", ReadOnly=1)
    ws_macro = wb_macro.Worksheets('価格アップロード')
    end_row = 0
    for file in files:

        time.sleep(1)

        fd = file.find('廃盤')
        if fd >= 0:
            continue

        #Excel開く
        s_book = px.load_workbook(file)
        #Excelブック内の各シートの名前をリストで取得できる
        s_sheet_name = s_book.sheetnames

        # #シートの総数を確認
        num_sheet = len(s_sheet_name)

        
        for i_sheet in s_book.sheetnames:
            seccuse = 0
            
            #前回のデータを削除
            if end_row != 0:
                arg = "B11:K" + str(end_row)
                r = ws_macro.Range(arg)
                ws_macro.Range(arg).Clear()
                # for row_index in range(1, end_row + 1):
                #     for col_index in range(1, r.Columns.Count + 1):
                #     row.append(r(row_index, col_index).Address)
                #     ret.append(row)
                # ws_macro.

          
            ws_s = s_book[i_sheet]
            product_cd = ws_s.cell(row=6, column=3).value
            if product_cd is None:
                continue
            #print('製品'+product_cd)    
            #最大行取得
            end_row = ws_s.max_row
            #print(str(end_row)+'行書き込み--開始--')   

            ws_macro.Cells.Item(5, 3).Value = '仕入価格のみ'
            ws_macro.Cells.Item(6, 3).value = product_cd
            ws_macro.Cells.Item(9, 5).value = '2020/01/01'
            ws_macro.Cells.Item(9, 6).value = '2020/01/01'

            wk = ws_macro.UsedRange
            cells = list(wk.value)

            # arg = "B11:G" + str(end_row)
            # ws_s_r = ws_macro.Range(arg)
            # ws_scells = ws_s_r.Value

            for rows in ws_s.iter_rows(min_row=11, min_col=2, max_row=end_row, max_col=7):
                for col in rows:
                    if col.value is None:
                        continue
                    ws_macro.Cells.Item(col.row,col.column).value =  col.value
                    # ws_scells[col.row-11][col.column-2] =  col.value

            #print(u'書き込み--完了--')        
            #print(u'名称取得--開始--')

            #価格Excel
            # ws_macro.Range(arg).value = ws_scells
            # ws_s_r = ws_s.Range(arg)
            # ws_scells = ws_s_r.Value
            # ws_scells = items


            main.bydupload(app,'nameSearch_click')
            
            if  ws_macro.Cells.Item(11, 8).value is None or  ws_macro.Cells.Item(11, 9).value is None:
                print('製品'+product_cd+'---名称取得失敗')
                continue
            else:
                #print('製品'+product_cd+'---名称取得完了')
                #print(u'アップロード--開始--')
                main.bydupload(app,'upload_click')
                #print(u'アップロード--完了--')
                
            arg = "J11:K" + str(end_row)
            r = ws_macro.Range(arg)
            cells = r.Value    
            flg = False
            for y in range(0,end_row-11):
                for x in range(0,1):
                    #print(cells[y][x])
                    if cells[y][x] == 'NG':
                        seccuse = 0
                        print('製品'+product_cd+'---' + str(cells[y][x+1]))
                        flg=True
                        break
                    else:
                        seccuse = 1
                if flg:
                    break
            if seccuse == 1:
                print('製品'+product_cd+'---アップロード完了')
    
               
        # if seccuse == 1:
        #     #print('製品'+product_cd+'---アップロード完了')      
        #     re_faile = 'E:\\src\\Python\\price_excel\\venv\\excel\済_'+ os.path.basename(file)
        #     os.rename(file,re_faile)
        # # else:
        # #     wb_macro.SageAs('E:\\src\\Python\\price_excel\\venv\\excel\\Excelツール\\TST【CH-MK】Excelツール_02.価格アップロード_'+ os.path.splitext(file)[0] +'.xlsm'+ )
    
    app.Quit()
    wb_macro.Close(SaveChanges=False)
    pythoncom.CoUninitialize()
    

# if  __name__ == "__main__":
    
#     files = glob.glob('E:\\src\\Python\\price_excel\\venv\\excel\*.xlsx')
#     byd_import(files)

